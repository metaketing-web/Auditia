"""Génération de pièces probantes démo pour la Data Room PNIPM."""
from __future__ import annotations

import csv
import io
import zipfile
from datetime import date
from typing import Any
from xml.sax.saxutils import escape

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

DEMO_BANNER = (
    "DOCUMENT DE DÉMONSTRATION — Mission d'audit PNIPM (MPJIPSC). "
    "Contenu fictif mais structuré pour les besoins de l'audit Skydeen."
)
MINISTRY = "Ministère de la Promotion de la Jeunesse, de l'Insertion Professionnelle et du Service Civique (MPJIPSC)"


def _pdf_escape(text: str) -> str:
    return text.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


def build_pdf(title: str, sections: list[tuple[str, list[str]]]) -> bytes:
    """PDF texte simple (une page ou plus) — encodage WinAnsi (accents FR)."""
    lines: list[str] = [title, "", DEMO_BANNER, ""]
    for heading, body in sections:
        lines.append(heading)
        lines.extend(body)
        lines.append("")
    text = "\n".join(lines).encode("latin-1", errors="replace").decode("latin-1")
    wrapped: list[str] = []
    for raw in text.split("\n"):
        while len(raw) > 92:
            wrapped.append(raw[:92])
            raw = raw[92:]
        wrapped.append(raw)

    y = 800
    stream_parts = ["BT", "/F1 11 Tf", "50 800 Td"]
    first = True
    for line in wrapped:
        if not first:
            stream_parts.append("0 -14 Td")
        first = False
        stream_parts.append(f"({_pdf_escape(line)}) Tj")
    stream_parts.append("ET")
    stream = "\n".join(stream_parts).encode("latin-1")
    objs: list[bytes] = []
    objs.append(b"1 0 obj<< /Type /Catalog /Pages 2 0 R >>endobj\n")
    objs.append(b"2 0 obj<< /Type /Pages /Kids [3 0 R] /Count 1 >>endobj\n")
    objs.append(
        b"3 0 obj<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] "
        b"/Contents 4 0 R /Resources << /Font << /F1 5 0 R >> >> >>endobj\n"
    )
    objs.append(
        f"4 0 obj<< /Length {len(stream)} >>stream\n".encode()
        + stream
        + b"\nendstream endobj\n"
    )
    objs.append(
        b"5 0 obj<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>endobj\n"
    )
    pdf = b"%PDF-1.4\n"
    offsets = [0]
    for obj in objs:
        offsets.append(len(pdf))
        pdf += obj
    xref_pos = len(pdf)
    pdf += f"xref\n0 {len(offsets)}\n0000000000 65535 f \n".encode()
    for off in offsets[1:]:
        pdf += f"{off:010d} 00000 n \n".encode()
    pdf += (
        f"trailer<< /Size {len(offsets)} /Root 1 0 R >>\n"
        f"startxref\n{xref_pos}\n%%EOF"
    ).encode()
    return pdf


def build_docx(title: str, paragraphs: list[str]) -> bytes:
    parts = [f"<w:p><w:r><w:rPr><w:b/></w:rPr><w:t>{escape(title)}</w:t></w:r></w:p>"]
    for p in paragraphs:
        parts.append(
            f"<w:p><w:r><w:t xml:space='preserve'>{escape(p)}</w:t></w:r></w:p>"
        )
    body = "".join(parts)
    document_xml = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">
  <w:body>{body}<w:sectPr/></w:body>
</w:document>"""
    content_types = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>
</Types>"""
    rels = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/>
</Relationships>"""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("[Content_Types].xml", content_types)
        zf.writestr("_rels/.rels", rels)
        zf.writestr("word/document.xml", document_xml)
    return buf.getvalue()


def _xlsx_header(ws, headers: list[str]) -> None:
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = fill
        c.font = font


def build_xlsx(sheet_name: str, headers: list[str], rows: list[list[Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    _xlsx_header(ws, headers)
    for r_idx, row in enumerate(rows, 2):
        for c_idx, val in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_csv(headers: list[str], rows: list[list[Any]]) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf, delimiter=";")
    w.writerow(headers)
    for row in rows:
        w.writerow(row)
    return buf.getvalue().encode("utf-8-sig")


def generate_document(rep: str, source: str, doc_type: str, desc: str, fmt: str) -> bytes:
    key = (rep, source, doc_type, desc, fmt)
    generators = {
        ("R2", "CABINET", "DOC", "fiches-poste-directions", "docx"): _fiches_poste,
        ("R3", "DPSD", "REF", "referentiel-donnees", "xlsx"): _referentiel_donnees,
        ("R3", "NNI", "DOC", "modalites-acces-nni", "pdf"): _modalites_nni,
        ("R4", "AEJ", "BASE", "base-beneficiaires-aej", "csv"): _base_aej,
        ("R4", "PEJEDEC", "BASE", "base-beneficiaires-pejedec", "csv"): _base_pejedec,
        ("R4", "USEP", "RAPPORT", "rapport-activite-usep-2025", "pdf"): _rapport_usep,
        ("R5", "DAF", "BUDGET", "budget-programmes-2026", "xlsx"): _budget_2026,
        ("R5", "DAF", "EXPORT", "export-sigfip-execution", "xlsx"): _export_sigfip,
        ("R5", "SIGFIP", "DOC", "modalites-raccordement-sigfip", "pdf"): _modalites_sigfip,
        ("R6", "DRH", "EFFECTIF", "etat-effectifs-ministere", "xlsx"): _effectifs_drh,
        ("R7", "DSI", "DOC", "cartographie-applicative", "pdf"): _carto_applicative,
        ("R7", "DSI", "DOC", "architecture-reseau-ipv6", "pdf"): _archi_reseau,
        ("R7", "UXP", "SPEC", "specs-raccordement-xroad", "pdf"): _specs_xroad,
        ("R8", "PATRI", "INVENTAIRE", "inventaire-flotte-automobile", "xlsx"): _inventaire_flotte,
        ("R9", "DAJC", "CONVENTION", "conventions-bailleurs", "pdf"): _conventions_bailleurs,
        ("R10", "BAD", "MODELE", "template-reporting-bad", "pdf"): _template_bad,
        ("R11", "DR-ABJ", "DOC", "procedures-saisie-regionales", "pdf"): _procedures_regionales,
    }
    fn = generators.get(key)
    if not fn:
        raise KeyError(f"Générateur absent : {key}")
    return fn()


def _fiches_poste() -> bytes:
    paras = [
        DEMO_BANNER,
        f"Périmètre : directions centrales et structures rattachées — {MINISTRY}",
        "Fiche n°1 — Directeur de Cabinet",
        "Mission : Assister le Ministre dans la coordination politique et le suivi des priorités gouvernementales.",
        "Compétences : pilotage inter-directions, préparation COPIL, interface Cabinet / programmes.",
        "Fiche n°2 — Directeur des Systèmes d'Information (DSI)",
        "Mission : Gouvernance du SI, sécurité, interopérabilité (UXP/X-Road), maintenance des applications métiers.",
        "Fiche n°3 — Directeur des Affaires Financières (DAF)",
        "Mission : Exécution budgétaire, liaison SIGFIP, contrôle des dépenses programmes.",
        "Fiche n°4 — Directeur de la Planification, Statistique et Documentation (DPSD)",
        "Mission : Consolidation des indicateurs, qualité des données, reporting vers le Cabinet et les bailleurs.",
        "Fiche n°5 — Directeur des Ressources Humaines (DRH)",
        "Mission : Gestion des effectifs, mobilité, formation, données RH pour le volet B.",
        "Version : v1 — Établi pour la mission d'audit PNIPM (juillet 2026).",
    ]
    return build_docx("Fiches de poste — Directions centrales MPJIPSC", paras)


def _referentiel_donnees() -> bytes:
    headers = ["Code indicateur", "Libellé", "Programme", "Périodicité", "Source", "Responsable", "Format", "Statut qualité"]
    rows = [
        ["IND-JEUN-01", "Jeunes inscrits (stock)", "Transversal", "Mensuel", "Registre AEJ", "DPSD", "Entier", "À fiabiliser"],
        ["IND-JEUN-02", "Jeunes insérés (flux)", "Transversal", "Mensuel", "Fichiers programmes", "DPSD", "Entier", "Doublons possibles"],
        ["IND-EMP-01", "Emplois créés USEP", "USEP", "Trimestriel", "UGP USEP", "Coord. USEP", "Entier", "Validé"],
        ["IND-SC-01", "Volontaires service civique", "OSCN", "Mensuel", "Plateforme OSCN", "OSCN", "Entier", "Partiel"],
        ["IND-FIN-01", "Taux exécution budgétaire", "Transversal", "Mensuel", "SIGFIP", "DAF", "Décimal", "Validé"],
        ["REF-TERR-01", "Code région", "Référentiel", "Statique", "Découpage admin.", "DPSD", "Texte", "Validé"],
        ["REF-TERR-02", "Code district / commune", "Référentiel", "Statique", "Découpage admin.", "DPSD", "Texte", "Incomplet"],
        ["REF-BEN-01", "Identifiant bénéficiaire interne", "Transversal", "—", "—", "DAJIP", "Texte", "Absent (pas de NNI)"],
    ]
    return build_xlsx("Referentiel", headers, rows)


def _modalites_nni() -> bytes:
    return build_pdf(
        "Modalités d'accès au NNI — Ministère de l'Intérieur",
        [
            ("1. Objet", [
                "Ce document décrit les conditions d'habilitation et d'appel aux services",
                "du Numéro National d'Identification (NNI) pour le MPJIPSC.",
            ]),
            ("2. Périmètre demandé", [
                "Vérification d'identité à l'inscription des bénéficiaires des dispositifs jeunesse.",
                "Pas de stockage du NNI en clair dans les applications actuelles du ministère.",
            ]),
            ("3. Procédure d'habilitation", [
                "Dossier à transmettre à la DGTCP / équipe NNI : acte de nomination du responsable de traitement,",
                "analyse d'impact, schéma d'architecture, liste des agents habilités.",
                "Délai indicatif : 8 à 12 semaines après dossier complet.",
            ]),
            ("4. État au " + date.today().isoformat(), [
                "Aucune habilitation effective pour le MPJIPSC à ce jour.",
                "Point focal désigné — entretien audit prévu S2.",
            ]),
        ],
    )


def _benef_rows(programme: str, prefix: str) -> list[list[Any]]:
    regions = ["Abidjan", "Abidjan", "Bouaké", "Korhogo", "San-Pédro", "Yamoussoukro"]
    rows = []
    for i in range(1, 26):
        rows.append([
            f"{prefix}-{i:04d}",
            programme,
            regions[i % len(regions)],
            "F" if i % 3 == 0 else "M",
            18 + (i % 12),
            f"2025-{(i % 12) + 1:02d}-{(i % 28) + 1:02d}",
            "Actif" if i % 5 else "Sorti",
        ])
    return rows


def _base_aej() -> bytes:
    return build_csv(
        ["id_interne", "programme", "region", "sexe", "age", "date_inscription", "statut"],
        _benef_rows("AEJ", "AEJ"),
    )


def _base_pejedec() -> bytes:
    return build_csv(
        ["id_interne", "programme", "region", "sexe", "age", "date_inscription", "statut"],
        _benef_rows("PEJEDEC", "PEJ"),
    )


def _rapport_usep() -> bytes:
    return build_pdf(
        "Rapport d'activité USEP — Exercice 2025",
        [
            ("Synthèse", [
                "Unité de Suivi-Évaluation et de Plaidoyer (USEP) — Programme PA-PSGouv / BAD.",
                "Période : janvier — décembre 2025. Version v1 transmise à l'audit PNIPM.",
            ]),
            ("Indicateurs clés 2025", [
                "Emplois temporaires créés (cible annuelle) : 12 400 — réalisé 9 870 (79,6 %).",
                "Taux de satisfaction bénéficiaires (enquête) : 72 %.",
                "Rapports trimestriels transmis au bailleur : 3/4 (Q4 en cours de consolidation).",
            ]),
            ("Points d'attention", [
                "Consolidation manuelle des données terrain (Excel régional).",
                "Absence de clé unique partagée avec les autres programmes du ministère.",
                "Recommandation interne : alignement sur le futur référentiel PNIPM.",
            ]),
        ],
    )


def _budget_2026() -> bytes:
    headers = ["Programme", "Action", "Ligne budgétaire", "AE 2026 (FCFA)", "CP notifiés", "Engagé", "Ordonnancé", "% exécution"]
    rows = [
        ["USEP", "A1", "1011", 4_500_000_000, 3_800_000_000, 2_100_000_000, 1_650_000_000, 43.4],
        ["PEJEDEC", "B2", "1023", 8_200_000_000, 7_000_000_000, 4_500_000_000, 3_200_000_000, 45.7],
        ["PACE", "C1", "1030", 2_100_000_000, 1_900_000_000, 890_000_000, 720_000_000, 37.9],
        ["Fonctionnement DSI", "—", "2015", 950_000_000, 950_000_000, 410_000_000, 380_000_000, 40.0],
        ["Fonctionnement Cabinet", "—", "2001", 620_000_000, 620_000_000, 280_000_000, 250_000_000, 40.3],
    ]
    return build_xlsx("Budget2026", headers, rows)


def _export_sigfip() -> bytes:
    headers = ["Exercice", "Ministère", "Programme", "Chapitre", "Engagement", "Liquidation", "Ordonnancement", "Date MAJ"]
    rows = [
        [2026, "MPJIPSC", "P001", "60", 2_100_000_000, 1_800_000_000, 1_650_000_000, "2026-06-30"],
        [2026, "MPJIPSC", "P002", "61", 4_500_000_000, 3_200_000_000, 3_000_000_000, "2026-06-30"],
        [2026, "MPJIPSC", "P003", "62", 890_000_000, 650_000_000, 620_000_000, "2026-06-28"],
    ]
    return build_xlsx("SIGFIP_export", headers, rows)


def _modalites_sigfip() -> bytes:
    return build_pdf(
        "Modalités de raccordement SIGFIP — MPJIPSC",
        [
            ("Contexte", [
                "Le SIGFIP (Système Intégré de Gestion des Finances Publiques) est opéré par le Ministère du Budget.",
                "Le MPJIPSC consulte les données d'exécution via export mensuel — pas d'interface temps réel.",
            ]),
            ("Modalités actuelles", [
                "Export Excel mensuel transmis par la DAF (délai J+15 après clôture mensuelle).",
                "Pas de webservice dédié. Pas de réconciliation automatique avec les indicateurs programmes.",
            ]),
            ("Besoin PNIPM", [
                "Flux automatisé ou API pour alimenter le tableau de bord consolidé.",
                "Traçabilité engagement → liquidation → ordonnancement → bénéficiaire (cible long terme).",
            ]),
        ],
    )


def _effectifs_drh() -> bytes:
    headers = ["Direction / Structure", "Catégorie A", "Catégorie B", "Catégorie C", "Contractuels", "Total", "Vacants"]
    rows = [
        ["Cabinet", 2, 8, 12, 4, 26, 1],
        ["DSI", 1, 6, 18, 9, 34, 2],
        ["DAF", 2, 10, 14, 3, 29, 0],
        ["DRH", 1, 5, 9, 2, 17, 1],
        ["DPSD", 1, 4, 11, 5, 21, 0],
        ["DAJIP", 2, 7, 15, 6, 30, 1],
        ["Patrimoine", 0, 3, 8, 4, 15, 1],
    ]
    return build_xlsx("Effectifs", headers, rows)


def _carto_applicative() -> bytes:
    return build_pdf(
        "Cartographie applicative — SI MPJIPSC (état des lieux)",
        [
            ("Socle actuel", [
                "12 applications identifiées — aucune architecture microservices.",
                "Registre AEJ (web), plateforme OSCN, fichiers Excel DPSD, outils comptables locaux DAF.",
            ]),
            ("Applications métier", [
                "AEJ-REG : gestion inscriptions Agence Emploi Jeunes.",
                "OSCN-PLATEFORME : suivi volontaires service civique.",
                "PROG-EXCEL : consolidation manuelle multi-programmes (DPSD).",
            ]),
            ("Intégrations", [
                "Aucun ESB / bus d'échange. Échanges par export CSV / e-mail.",
                "Pas de raccordement UXP/X-Road documenté.",
            ]),
            ("Cible PNIPM (CDC)", [
                "Socle unique, API ouvertes, référentiel données partagé, traçabilité bout-en-bout.",
            ]),
        ],
    )


def _archi_reseau() -> bytes:
    return build_pdf(
        "Architecture réseau & préparation IPv6 — DSI",
        [
            ("Infrastructure", [
                "Datacenter ministériel mutualisé — virtualisation VMware.",
                "Liaisons MPLS vers 3 directions régionales pilotes (Abidjan, Bouaké, Korhogo).",
            ]),
            ("IPv4 / IPv6", [
                "IPv4 : plan d'adressage 10.20.0.0/16 — saturation prévue sous 24 mois.",
                "IPv6 : plan d'adressage 2001:db8:mpji::/48 proposé — déploiement non engagé.",
                "Dual-stack non activé sur le pare-feu périphérique.",
            ]),
            ("Sécurité", [
                "Pare-feu Fortinet — règles revues annuellement.",
                "Segmentation VLAN partielle (prod / admin / invités).",
                "SOC externalisé — contrat en renégociation.",
            ]),
        ],
    )


def _specs_xroad() -> bytes:
    return build_pdf(
        "Spécifications de raccordement X-Road / UXP — Côte d'Ivoire",
        [
            ("Référentiel", [
                "UXP (eXtended Unified Platform) — équivalent national X-Road.",
                "Autorité de certification : ANSUTI / équipe gouvernance interopérabilité.",
            ]),
            ("Prérequis raccordement", [
                "1. Création instance sécurisée (serveur membre X-Road).",
                "2. Certificats d'authentification et de signature.",
                "3. Publication des services (REST/SOAP) dans le catalogue national.",
                "4. Tests en environnement de qualification.",
            ]),
            ("État MPJIPSC", [
                "Aucune instance déployée. Dossier d'intention non déposé.",
                "Services cibles identifiés : vérification NNI, consultation SIGFIP (lecture).",
            ]),
        ],
    )


def _inventaire_flotte() -> bytes:
    headers = ["Immatriculation", "Marque / Modèle", "Année", "Affectation", "État", "Kilométrage", "Assurance", "Observations"]
    rows = [
        ["AB-1234-CI", "Toyota Hilux", 2019, "Cabinet", "Bon", 87_400, "Valide", "—"],
        ["AB-5678-CI", "Mitsubishi L200", 2018, "DSI", "Moyen", 112_000, "Valide", "Révision due"],
        ["BK-9012-CI", "Nissan Navara", 2020, "DR Abidjan", "Bon", 65_200, "Valide", "Terrain S4"],
        ["KO-3456-CI", "Toyota Land Cruiser", 2017, "DR Nord", "Moyen", 145_800, "Valide", "—"],
        ["SP-7890-CI", "Hyundai H1", 2021, "Patrimoine", "Bon", 42_100, "Valide", "Pool ministériel"],
    ]
    return build_xlsx("Flotte", headers, rows)


def _conventions_bailleurs() -> bytes:
    return build_pdf(
        "Conventions de financement — Synthèse bailleurs (MPJIPSC)",
        [
            ("BAD", [
                "PACE / USEP — Convention Côte d'Ivoire PACE-CI-2022.",
                "Reporting trimestriel — indicateurs emploi jeunes, genre, inclusion.",
            ]),
            ("Banque mondiale", [
                "PEJEDEC — Financement IDA, composante emploi des jeunes.",
                "Clauses de décaissement liées aux indicateurs de performance (DLI).",
            ]),
            ("AFD", [
                "C2D-EMPLOI — Appui à l'insertion professionnelle.",
                "Rapports semestriels — format Excel + note narrative.",
            ]),
            ("PNUD / BOAD", [
                "PSIE Innovation, UGP-CSC — conventions signées 2023-2024.",
                "Obligations de transparence et de publication open data (partielle).",
            ]),
            ("Point d'attention audit", [
                "Hétérogénéité des formats de reporting — consolidation manuelle au Cabinet.",
            ]),
        ],
    )


def _template_bad() -> bytes:
    return build_pdf(
        "Template reporting BAD — Indicateurs trimestriels PACE/USEP",
        [
            ("Structure du rapport", [
                "Section A — Données générales (période, UGP, contact).",
                "Section B — Indicateurs quantitatifs (cibles / réalisations / écarts).",
                "Section C — Indicateurs qualitatifs et études de cas.",
                "Section D — Contraintes, risques, demandes d'arbitrage.",
            ]),
            ("Indicateurs type", [
                "EMP-T1 : Emplois créés (désagrégation H/F, handicap, région).",
                "EMP-T2 : Taux de rétention à 6 mois.",
                "GEN-T1 : Part femmes parmi bénéficiaires (cible 40 %).",
            ]),
            ("Instructions", [
                "Remplir une ligne par district. Joindre liste bénéficiaires anonymisée.",
                "Délai de transmission : J+30 après fin de trimestre.",
            ]),
        ],
    )


def _procedures_regionales() -> bytes:
    return build_pdf(
        "Procédures de saisie régionale — Direction Régionale Abidjan",
        [
            ("Objectif", [
                "Harmoniser la collecte terrain des données programmes avant consolidation DPSD.",
            ]),
            ("Workflow", [
                "1. Agent régional saisit fiche bénéficiaire (formulaire Excel local).",
                "2. Chef de service valide hebdomadairement.",
                "3. Envoi e-mail chiffré au référent DPSD (vendredi 17h).",
                "4. Contrôle qualité DPSD — retour corrections sous 72h.",
            ]),
            ("Règles de saisie", [
                "Champs obligatoires : nom, prénom, date naissance, sexe, commune, programme, date inscription.",
                "Pas de NNI systématique — doublons gérés manuellement.",
            ]),
            ("Écarts identifiés", [
                "3 versions de formulaire en circulation (Abidjan, Centre, Nord).",
                "Délais de consolidation : 5 à 12 jours selon la région.",
            ]),
        ],
    )


def normalized_filename(rep: str, source: str, doc_type: str, desc: str, version: str, fmt: str) -> str:
    return f"prevu_{rep}_{source}_{doc_type}_{desc}_{version}.{fmt}"


def mime_for(fmt: str) -> str:
    return {
        "pdf": "application/pdf",
        "docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "csv": "text/csv; charset=utf-8",
    }.get(fmt, "application/octet-stream")
